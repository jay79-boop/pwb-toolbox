"""Build the Profit & Exit Planner workbook.

This replaces a hand-maintained Google Sheet ("Profit Gain Percentage") in which
every number was typed rather than computed: prices frozen at their October 2022
values, positions duplicated across rows, an exit ladder that sold more coins
than the position held, and a 1,900-cell percentage grid that was a single
formula repeated by hand.

Everything here is generated, so the workbook can be rebuilt from scratch
instead of patched in place:

    python tools/build_profit_planner.py --out profit_planner.xlsx

Upload the result to Google Drive and it converts to a Google Sheet with the
formulas live. Price cells call GOOGLEFINANCE and fall back to a manually
entered price when the feed does not carry the symbol, so the sheet degrades to
"still correct, just stale" rather than to "#N/A everywhere".
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------
# Palette. Ink for structure, amber for "you type here", grey for "pulled from
# somewhere else". The legend on Start Here documents this contract; keep the
# two in step.
# --------------------------------------------------------------------------

INK = "0F172A"
SLATE = "334155"
MUTED = "64748B"
LINE = "CBD5E1"
BAND = "F8FAFC"
DERIVED = "F1F5F9"
INPUT_BG = "FEF3C7"
INPUT_LINE = "F59E0B"
ACCENT = "0D9488"
ACCENT_SOFT = "CCFBF1"
GOOD = "15803D"
GOOD_SOFT = "DCFCE7"
BAD = "B91C1C"
BAD_SOFT = "FEE2E2"
WARN_SOFT = "FEF9C3"
WHITE = "FFFFFF"

MONEY = '"$"#,##0.00'
MONEY0 = '"$"#,##0'
MONEY_FINE = '"$"#,##0.00000000'
QTY = "#,##0.########"
PCT = "0.0%"
PCT0 = "0%"
MULT = '0.00"x"'

TITLE_F = Font(name="Calibri", size=16, bold=True, color=WHITE)
SUB_F = Font(name="Calibri", size=10, italic=True, color=MUTED)
HEAD_F = Font(name="Calibri", size=10, bold=True, color=WHITE)
BODY_F = Font(name="Calibri", size=11, color=INK)
BOLD_F = Font(name="Calibri", size=11, bold=True, color=INK)
SMALL_F = Font(name="Calibri", size=9, color=MUTED)
DERIVED_F = Font(name="Calibri", size=11, italic=True, color=SLATE)
KPI_F = Font(name="Calibri", size=20, bold=True, color=INK)
KPI_LABEL_F = Font(name="Calibri", size=9, bold=True, color=MUTED)

TITLE_FILL = PatternFill("solid", fgColor=INK)
HEAD_FILL = PatternFill("solid", fgColor=SLATE)
INPUT_FILL = PatternFill("solid", fgColor=INPUT_BG)
DERIVED_FILL = PatternFill("solid", fgColor=DERIVED)
BAND_FILL = PatternFill("solid", fgColor=BAND)
ACCENT_FILL = PatternFill("solid", fgColor=ACCENT_SOFT)

THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
INPUT_SIDE = Side(style="thin", color=INPUT_LINE)
INPUT_BORDER = Border(
    left=INPUT_SIDE, right=INPUT_SIDE, top=INPUT_SIDE, bottom=INPUT_SIDE
)

# Row geography of the Positions sheet. Several other sheets index into it, so
# it is stated once here rather than spelled out at each call site.
POS_FIRST = 4
POS_LAST = 33
POS_TOTAL = 35

LOG_FIRST = 5
LOG_LAST = 204

# Row geography of a Plan tab. These are the tabs that get used daily,
# so the layout is fixed here and every formula on it refers back to these.
# The position reads across one strip (labels on TP_LABEL, values on TP_INPUT)
# rather than down a column, which is what keeps the ladder above the fold.
TP_LABEL = 4
TP_INPUT = 5
TP_TILE_LABEL = 7
TP_TILE_VALUE = 8
TP_BANNER = 11
TP_HEADER = 12
TP_RUNG_FIRST = 13
TP_RUNG_LAST = 26

# Six plans ship ready-made so the Dashboard can check them. A seventh coin is
# a duplicated tab, which works but sits outside the check.
PLAN_COUNT = 6

# (gain above your average cost, share of what is still held at that rung).
# Shares of the remainder rather than of the original position, so the plan
# cannot oversell and every rung re-sizes when one above it changes. They
# compound to roughly 90% sold, leaving a tenth of the position riding.
TP_RUNGS = [
    (0.05, 0.10),
    (0.10, 0.10),
    (0.15, 0.15),
    (0.20, 0.15),
    (0.25, 0.20),
    (0.30, 0.20),
    (0.40, 0.25),
    (0.50, 0.25),
    (0.75, 0.30),
    (1.00, 0.33),
]


@dataclass
class Position:
    """A holding carried over from the old sheet.

    ``manual_price`` is the last price the old sheet recorded. It is both the
    seed value and the fallback the price formula lands on when GOOGLEFINANCE
    does not carry the symbol.
    """

    asset: str
    ticker: str
    sector: str
    qty: float
    avg_cost: float
    manual_price: float
    ath: float | None
    note: str = ""
    asset_class: str = "Crypto"


# --------------------------------------------------------------------------
# Carried-over data.
#
# Sourced from the old workbook's "Crypto Gaines" tab, which was the only tab
# that read as a holdings list rather than a plan. Duplicate rows (Decentraland,
# Chainlink) and the two "Future Asset" placeholders are dropped. Average cost is
# the old Cost column divided by the old Amount column, so it survives as one
# number instead of two that could disagree.
# --------------------------------------------------------------------------

# Observed from a live import: Google Finance carries the majors as
# CURRENCY:XXXUSD and nothing else. Bitcoin, Ethereum, Cardano and XRP resolve;
# these do not, and a row that cannot be priced sits at whatever was typed into
# it — which is how a portfolio total quietly becomes fiction.
NO_GOOGLE_FEED = {
    "Chainlink",
    "Gala",
    "The Sandbox",
    "Decentraland",
    "Theta",
    "Harmony",
    "Audius",
    "Shiba Inu",
}

POSITIONS: list[Position] = [
    Position(
        "EXAMPLE — overwrite or delete this row",
        "CURRENCY:BTCUSD",
        "Store of value",
        0.05,
        50000,
        50000,
        126000,
        "Every column filled in, so you can see the shape of a complete "
        "holding. Overwrite it with your first real position or delete the row.",
    ),
    Position(
        "Bitcoin",
        "CURRENCY:BTCUSD",
        "Store of value",
        0,
        0,
        0,
        None,
        "Google Finance carries this one. Type your quantity and average cost "
        "and the price is already live.",
    ),
    Position(
        "Ethereum",
        "CURRENCY:ETHUSD",
        "L1 / Smart contract",
        0,
        0,
        0,
        None,
        "Google Finance carries this one. Type your quantity and average cost "
        "and the price is already live.",
    ),
]

# The workbook is one book for everything, not one per asset type: a stock and
# a coin differ only in which ticker the price feed wants.
CLASSES = ["Crypto", "Stock", "ETF", "Option", "Cash", "Other"]

# Closed keeps a finished round on the register with its realised result and
# out of every portfolio total; Watchlist is a position sized at zero.
STATUSES = ["Active", "Closed", "Watchlist"]

SECTORS = [
    "Store of value",
    "L1 / Smart contract",
    "L2 / Scaling",
    "DeFi / Oracle",
    "Gaming",
    "Metaverse",
    "Media / NFT",
    "Exchange",
    "Low cap",
    "Meme",
    "Equity",
    "Review",
    "Other",
]


# --------------------------------------------------------------------------
# Small styling helpers.
# --------------------------------------------------------------------------


def sheet_title(ws, text: str, subtitle: str, span: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = TITLE_F
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    sub = ws.cell(row=2, column=1, value=subtitle)
    sub.font = SUB_F
    sub.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 18


def header_row(ws, row: int, headers: list[str], widths: list[int]) -> None:
    for idx, (label, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=idx, value=label)
        cell.font = HEAD_F
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        cell.border = BOX
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[row].height = 30


def style_input(cell) -> None:
    cell.fill = INPUT_FILL
    cell.border = INPUT_BORDER
    cell.font = BODY_F


def style_derived(cell) -> None:
    cell.fill = DERIVED_FILL
    cell.border = BOX
    cell.font = DERIVED_F


def style_body(cell) -> None:
    cell.border = BOX
    cell.font = BODY_F


def label_value(ws, row: int, label: str, value, fmt=None, note: str = ""):
    """One `label | value | note` line, with the value marked as an input."""
    lab = ws.cell(row=row, column=2, value=label)
    lab.font = BOLD_F
    cell = ws.cell(row=row, column=3, value=value)
    style_input(cell)
    if fmt:
        cell.number_format = fmt
    if note:
        hint = ws.cell(row=row, column=4, value=note)
        hint.font = SMALL_F
        hint.alignment = Alignment(vertical="center", wrap_text=True)
    return cell


def pos_lookup(column: str, key_cell: str) -> str:
    """INDEX/MATCH into Positions.

    Deliberately not XLOOKUP: this workbook is written as .xlsx and imported,
    and INDEX/MATCH survives that trip everywhere.
    """
    return (
        f"INDEX(Positions!${column}${POS_FIRST}:${column}${POS_LAST},"
        f"MATCH({key_cell},Positions!$A${POS_FIRST}:$A${POS_LAST},0))"
    )


# --------------------------------------------------------------------------
# Sheets.
# --------------------------------------------------------------------------


def build_lists(wb: Workbook):
    ws = wb.create_sheet("Lists")
    ws.sheet_state = "hidden"
    for col, (header, values) in enumerate(
        [
            ("Sector", SECTORS),
            ("Price mode", ["Auto", "Manual"]),
            ("Side", ["Buy", "Sell"]),
            ("Yes/No", ["Yes", "No"]),
            ("Class", CLASSES),
            ("Status", STATUSES),
        ],
        start=1,
    ):
        ws.cell(row=1, column=col, value=header)
        for offset, value in enumerate(values, start=2):
            ws.cell(row=offset, column=col, value=value)
    return ws


def build_settings(wb: Workbook):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = MUTED
    ws.sheet_view.showGridLines = False
    sheet_title(
        ws,
        "Settings",
        "Four numbers the rest of the workbook reads. Change them here and "
        "every tab follows.",
        8,
    )
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 78

    label_value(
        ws,
        4,
        "Capital gains tax rate",
        0.24,
        PCT,
        "Used to show every exit after tax, not before. A guess is fine — a "
        "wrong rate here is still closer than ignoring tax entirely.",
    )
    label_value(
        ws,
        5,
        "Exchange / broker fee",
        0.005,
        "0.000%",
        "Charged on each sale in the ladder and calculator. 0.5% is Coinbase's "
        "rough taker fee; a broker with free equity trades is 0%.",
    )
    label_value(
        ws,
        6,
        "Max weight per position",
        0.20,
        PCT,
        "Positions above this are flagged amber on the Positions tab and "
        "counted on the Dashboard. It is a tripwire, not a rule.",
    )
    label_value(
        ws,
        7,
        "Portfolio name",
        "Main book",
        None,
        "Shown on the Dashboard.",
    )

    ws["B9"] = "Colour code"
    ws["B9"].font = BOLD_F
    swatches = [
        (
            INPUT_BG,
            INPUT_LINE,
            "Amber = type here. Nothing else expects to be typed in.",
        ),
        (
            DERIVED,
            LINE,
            "Grey italic = pulled from another tab. Overwriting it breaks the link.",
        ),
        (WHITE, LINE, "White = calculated. Leave it alone and it stays right."),
    ]
    for offset, (bg, edge, text) in enumerate(swatches):
        row = 10 + offset
        chip = ws.cell(row=row, column=2, value="")
        chip.fill = PatternFill("solid", fgColor=bg)
        side = Side(style="thin", color=edge)
        chip.border = Border(left=side, right=side, top=side, bottom=side)
        desc = ws.cell(row=row, column=3, value=text)
        desc.font = SMALL_F
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text="Settings!$C$4")
    wb.defined_names["FeeRate"] = DefinedName("FeeRate", attr_text="Settings!$C$5")
    wb.defined_names["MaxWeight"] = DefinedName("MaxWeight", attr_text="Settings!$C$6")
    return ws


def build_positions(wb: Workbook):
    ws = wb.create_sheet("Positions")
    ws.sheet_properties.tabColor = ACCENT
    sheet_title(
        ws,
        "Positions",
        "The register: everything you hold, everything you used to hold, and "
        "anything you are watching. Closed rows keep their realised result and "
        "stop counting toward the portfolio.",
        24,
    )
    headers = [
        "Asset",
        "Ticker",
        "Class",
        "Sector",
        "Status",
        "Round",
        "Quantity",
        "Avg cost",
        "Cost basis",
        "Price mode",
        "Manual price",
        "Price",
        "Feed",
        "Market value",
        "Unrealised $",
        "Unrealised %",
        "Weight",
        "Realised $",
        "Total result",
        "All-time high",
        "% of ATH",
        "X to ATH",
        "If ATH returns",
        "Note",
        "Needs",
    ]
    widths = [
        28,
        19,
        10,
        18,
        10,
        8,
        14,
        13,
        13,
        11,
        13,
        13,
        12,
        14,
        13,
        12,
        9,
        13,
        13,
        13,
        9,
        9,
        14,
        58,
    ]
    header_row(ws, 3, headers, widths)
    ws.freeze_panes = "C4"

    dropdowns = {
        3: "=Lists!$E$2:$E$7",
        4: "=Lists!$A$2:$A$14",
        5: "=Lists!$F$2:$F$4",
        10: "=Lists!$B$2:$B$3",
    }
    validators = {}
    for col, source in dropdowns.items():
        dv = DataValidation(type="list", formula1=source, allow_blank=True)
        ws.add_data_validation(dv)
        validators[col] = dv

    for row in range(POS_FIRST, POS_LAST + 1):
        seed = POSITIONS[row - POS_FIRST] if row - POS_FIRST < len(POSITIONS) else None
        banded = (row - POS_FIRST) % 2 == 1

        for col in range(1, 26):
            cell = ws.cell(row=row, column=col)
            style_body(cell)
            if banded:
                cell.fill = BAND_FILL
        for col in (1, 2, 3, 4, 5, 6, 7, 8, 10, 11, 20, 24):
            style_input(ws.cell(row=row, column=col))

        fine = seed is not None and seed.manual_price < 0.01
        ws.cell(row=row, column=1, value=seed.asset if seed else None)
        ws.cell(row=row, column=2, value=seed.ticker if seed else None)
        ws.cell(row=row, column=3, value=seed.asset_class if seed else None)
        ws.cell(row=row, column=4, value=seed.sector if seed else None)
        ws.cell(row=row, column=5, value="Active" if seed else None)
        ws.cell(row=row, column=6, value=1 if seed else None).alignment = Alignment(
            horizontal="center"
        )
        qty = ws.cell(row=row, column=7, value=seed.qty if seed else None)
        qty.number_format = QTY
        avg = ws.cell(row=row, column=8, value=seed.avg_cost if seed else None)
        avg.number_format = MONEY_FINE if fine else MONEY
        ws.cell(row=row, column=10, value="Auto" if seed else None)
        manual = ws.cell(row=row, column=11, value=seed.manual_price if seed else None)
        manual.number_format = MONEY_FINE if fine else MONEY
        ath = ws.cell(row=row, column=20, value=seed.ath if seed else None)
        ath.number_format = MONEY_FINE if fine else MONEY
        note = ws.cell(row=row, column=24, value=seed.note if seed else None)
        note.font = SMALL_F
        note.alignment = Alignment(vertical="top", wrap_text=True)

        for col, dv in validators.items():
            dv.add(ws.cell(row=row, column=col))

        # Guarded on the name, not the quantity: a closed position keeps its
        # quantity as a record of what it was, and still has to compute a
        # realised result.
        blank = f'$A{row}=""'
        inactive = f'OR($A{row}="",$E{row}<>"Active")'
        ws.cell(
            row=row, column=9, value=f'=IF({blank},"",$G{row}*$H{row})'
        ).number_format = MONEY
        ws.cell(
            row=row,
            column=12,
            value=(
                f'=IF({blank},"",IFERROR(IF($J{row}="Manual",$K{row},'
                f"GOOGLEFINANCE($B{row})),$K{row}))"
            ),
        ).number_format = (
            MONEY_FINE if fine else MONEY
        )
        ws.cell(
            row=row,
            column=13,
            value=(
                f'=IF({blank},"",IF($J{row}="Manual","manual",'
                f'IF(ISERROR(GOOGLEFINANCE($B{row})),"no feed","live")))'
            ),
        ).font = SMALL_F
        # Everything from here to Weight is deliberately blank unless the
        # position is Active, which is what keeps a closed round out of the
        # portfolio totals without deleting it.
        ws.cell(
            row=row, column=14, value=f'=IF({inactive},"",$G{row}*$L{row})'
        ).number_format = MONEY
        ws.cell(
            row=row, column=15, value=f'=IF($N{row}="","",$N{row}-$I{row})'
        ).number_format = MONEY
        ws.cell(
            row=row,
            column=16,
            value=f'=IF(OR($O{row}="",$I{row}=0),"",$O{row}/$I{row})',
        ).number_format = PCT
        ws.cell(
            row=row,
            column=17,
            value=f'=IF(OR($N{row}="",$N${POS_TOTAL}=0),"",$N{row}/$N${POS_TOTAL})',
        ).number_format = PCT
        ws.cell(
            row=row,
            column=18,
            value=(
                f"=IF({blank},\"\",SUMIFS('Trade Log'!$L${LOG_FIRST}:$L${LOG_LAST},"
                f"'Trade Log'!$B${LOG_FIRST}:$B${LOG_LAST},$A{row},"
                f"'Trade Log'!$C${LOG_FIRST}:$C${LOG_LAST},$F{row}))"
            ),
        ).number_format = MONEY
        ws.cell(
            row=row, column=19, value=f'=IF({blank},"",N($O{row})+N($R{row}))'
        ).number_format = MONEY
        ws.cell(
            row=row,
            column=21,
            value=f'=IF(OR({blank},$T{row}=""),"",$L{row}/$T{row})',
        ).number_format = PCT
        ws.cell(
            row=row,
            column=22,
            value=f'=IF(OR({blank},$T{row}="",$L{row}=0),"",$T{row}/$L{row})',
        ).number_format = MULT
        ws.cell(
            row=row,
            column=23,
            value=f'=IF(OR({inactive},$T{row}=""),"",$G{row}*$T{row})',
        ).number_format = MONEY
        # A named row that is missing something has to say so. An earlier
        # version let a holding with no Status sit at zero, contributing
        # nothing and reporting nothing — the register simply ignored it.
        missing = ws.cell(
            row=row,
            column=25,
            value=(
                f'=IF({blank},"",'
                f'IF($E{row}="","set Status",'
                f'IF($G{row}="","set Quantity",'
                f'IF($H{row}="","set Avg cost",'
                f'IF($C{row}="","set Class",'
                f'IF($B{row}="","set Ticker",""))))))'
            ),
        )
        missing.font = Font(name="Calibri", size=9, bold=True, color="92400E")

    for col in range(1, 26):
        cell = ws.cell(row=POS_TOTAL, column=col)
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.border = BOX
    total = ws.cell(
        row=POS_TOTAL, column=1, value="TOTAL — active only, except realised"
    )
    total.alignment = Alignment(indent=1)
    ws.merge_cells(start_row=POS_TOTAL, start_column=1, end_row=POS_TOTAL, end_column=6)
    ws.cell(
        row=POS_TOTAL,
        column=9,
        value=f'=SUMIFS(I{POS_FIRST}:I{POS_LAST},E{POS_FIRST}:E{POS_LAST},"Active")',
    ).number_format = MONEY
    for col, letter in ((14, "N"), (15, "O"), (18, "R"), (19, "S"), (23, "W")):
        ws.cell(
            row=POS_TOTAL,
            column=col,
            value=f"=SUM({letter}{POS_FIRST}:{letter}{POS_LAST})",
        ).number_format = MONEY
    ws.cell(
        row=POS_TOTAL,
        column=16,
        value=f'=IF($I${POS_TOTAL}=0,"",$O${POS_TOTAL}/$I${POS_TOTAL})',
    ).number_format = PCT

    for span in (f"O{POS_FIRST}:P{POS_LAST}", f"R{POS_FIRST}:S{POS_LAST}"):
        ws.conditional_formatting.add(
            span,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                font=Font(color=BAD),
                fill=PatternFill("solid", fgColor=BAD_SOFT),
            ),
        )
        ws.conditional_formatting.add(
            span,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                font=Font(color=GOOD),
                fill=PatternFill("solid", fgColor=GOOD_SOFT),
            ),
        )
    ws.conditional_formatting.add(
        f"Q{POS_FIRST}:Q{POS_LAST}",
        DataBarRule(
            start_type="num", start_value=0, end_type="num", end_value=1, color=ACCENT
        ),
    )
    ws.conditional_formatting.add(
        f"Q{POS_FIRST}:Q{POS_LAST}",
        FormulaRule(
            formula=[f'AND($Q{POS_FIRST}<>"",$Q{POS_FIRST}>MaxWeight)'],
            fill=PatternFill("solid", fgColor=WARN_SOFT),
        ),
    )
    ws.conditional_formatting.add(
        f"M{POS_FIRST}:M{POS_LAST}",
        CellIsRule(
            operator="equal", formula=['"no feed"'], font=Font(color=BAD, size=9)
        ),
    )
    ws.conditional_formatting.add(
        f"A{POS_FIRST}:Y{POS_LAST}",
        FormulaRule(formula=[f'$E{POS_FIRST}="Closed"'], font=Font(color=MUTED)),
    )
    ws.conditional_formatting.add(
        f"A{POS_FIRST}:Y{POS_LAST}",
        FormulaRule(
            formula=[f'AND($A{POS_FIRST}<>"",$Y{POS_FIRST}<>"")'],
            fill=PatternFill("solid", fgColor=WARN_SOFT),
        ),
    )
    ws.conditional_formatting.add(
        f"E{POS_FIRST}:E{POS_LAST}",
        CellIsRule(
            operator="equal",
            formula=['"Watchlist"'],
            font=Font(color="1D4ED8", bold=True),
        ),
    )
    return ws


def build_plan_tab(wb: Workbook, index: int):
    """One coin, one screen, the whole exit plan.

    Six of these ship ready-made. Each picks a holding from the register by
    name, so quantity, cost and ticker arrive rather than being typed twice —
    the drift that put XRP in the workbook at two different average costs
    cannot happen by accident. Those cells stay amber: typing over one is a
    legitimate what-if, and the tab then says it no longer matches the
    register, which is a warning that means something specific.

    The only sheet a plan refers to is Positions, so duplicating a tab for a
    seventh coin still works — the copy picks its own holding and computes from
    that, rather than mirroring the tab it came from.
    """
    ws = wb.create_sheet(f"Plan {index}")
    ws.sheet_properties.tabColor = ACCENT
    ws.sheet_view.showGridLines = False
    sheet_title(
        ws,
        f"Plan {index}",
        "Click the first amber cell and choose one of your holdings from the "
        "list. Its ticker, quantity and average cost fill themselves in. If "
        "the coin you want is not in the list, add it on the Positions tab "
        "first. Duplicate this tab for a seventh coin.",
        20,
    )

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 20
    for col in "CDEFGHIJKLMNO":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["P"].width = 2
    for col in "QRST":
        ws.column_dimensions[col].width = 15

    row = TP_INPUT
    holding = f"$B${row}"
    ticker, units, cost = f"$C${row}", f"$D${row}", f"$E${row}"
    mode, yours, live = f"$F${row}", f"$G${row}", f"$H${row}"
    basis = f"$J${row}"

    def from_register(column: str) -> str:
        return f'IFERROR({pos_lookup(column, holding)},"")'

    strip = [
        # Plan 1 opens already pointing at the example row: a dropdown nobody
        # has used yet does not explain what it wants.
        (
            2,
            "Pick from Positions",
            POSITIONS[0].asset if index == 1 else None,
            None,
            True,
        ),
        (3, "Ticker", f'=IF({holding}="","",{from_register("B")})', None, False),
        (4, "Units held", f'=IF({holding}="","",{from_register("G")})', QTY, True),
        (5, "Average cost", f'=IF({holding}="","",{from_register("H")})', MONEY, True),
        (6, "Price mode", "Auto", None, True),
        (7, "Your price", f'=IF({holding}="","",{from_register("K")})', MONEY, True),
        (
            8,
            "Price used",
            f'=IF({mode}="Manual",{yours},IFERROR(GOOGLEFINANCE({ticker}),{yours}))',
            MONEY,
            False,
        ),
        (
            9,
            "Feed",
            f'=IF({mode}="Manual","manual — pinned by you",'
            f'IF(ISERROR(GOOGLEFINANCE({ticker})),"no feed — using your price","live"))',
            None,
            False,
        ),
        (10, "Cost basis", f"={units}*{cost}", MONEY, False),
        (11, "Market value", f"={units}*{live}", MONEY, False),
        (12, "Open profit", f"=$K${row}-{basis}", MONEY, False),
        (13, "Open profit %", f'=IF({basis}=0,"",$L${row}/{basis})', PCT, False),
        # Typing over a pulled cell is allowed, but the tab has to admit it is
        # no longer showing what the register holds.
        (
            14,
            "Matches register",
            (
                # A name that is not on the register produced blanks all the
                # way down and still reported a match, because two empty cells
                # compare equal. Say it plainly instead.
                f'=IF({holding}="","",'
                f"IF(COUNTIF(Positions!$A${POS_FIRST}:$A${POS_LAST},{holding})=0,"
                f'"NOT ON THE REGISTER — add it on Positions first",'
                f"IF(AND(ROUND(N({units}),8)=ROUND(N({from_register('G')}),8),"
                f"ROUND(N({cost}),8)=ROUND(N({from_register('H')}),8)),"
                f'"yes","EDITED")))'
            ),
            None,
            False,
        ),
    ]
    for col, label, value, fmt, typed in strip:
        head = ws.cell(row=TP_LABEL, column=col, value=label.upper())
        head.font = KPI_LABEL_F
        head.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        head.fill = PatternFill("solid", fgColor=INPUT_BG if typed else DERIVED)
        head.border = BOX
        cell = ws.cell(row=TP_INPUT, column=col, value=value)
        style_input(cell) if typed else style_derived(cell)
        if fmt:
            cell.number_format = fmt
    ws.row_dimensions[TP_LABEL].height = 26
    ws.row_dimensions[TP_INPUT].height = 20

    holding_dv = DataValidation(
        type="list",
        formula1=f"=Positions!$A${POS_FIRST}:$A${POS_LAST}",
        allow_blank=True,
        promptTitle="Choose one of your holdings",
        prompt=(
            "This is a list of the names on your Positions tab — Bitcoin, "
            "Ethereum, and anything else you have added there. Choose one and "
            "this plan reads its ticker, quantity and average cost.\n\n"
            "Not a coin type, and not a ticker symbol. If what you want is "
            "not listed, add a row for it on Positions first."
        ),
        showInputMessage=True,
    )
    mode_dv = DataValidation(type="list", formula1='"Auto,Manual"', allow_blank=True)
    for dv, column in ((holding_dv, 2), (mode_dv, 6)):
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=TP_INPUT, column=column))

    for ref in (f"L{TP_INPUT}", f"M{TP_INPUT}"):
        ws.conditional_formatting.add(
            ref,
            CellIsRule(
                operator="lessThan", formula=["0"], font=Font(bold=True, color=BAD)
            ),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(
                operator="greaterThan", formula=["0"], font=Font(bold=True, color=GOOD)
            ),
        )
    ws.conditional_formatting.add(
        f"N{TP_INPUT}",
        CellIsRule(
            operator="equal",
            formula=['"EDITED"'],
            font=Font(bold=True, color="92400E"),
            fill=PatternFill("solid", fgColor=WARN_SOFT),
        ),
    )
    ws.conditional_formatting.add(
        f"N{TP_INPUT}",
        FormulaRule(
            formula=[f'LEFT($N${TP_INPUT},6)="NOT ON"'],
            font=Font(bold=True, color=BAD),
            fill=PatternFill("solid", fgColor=BAD_SOFT),
        ),
    )
    ws.conditional_formatting.add(
        f"I{TP_INPUT}",
        CellIsRule(
            operator="equal",
            formula=['"manual — pinned by you"'],
            font=Font(bold=True, color="92400E"),
            fill=PatternFill("solid", fgColor=WARN_SOFT),
        ),
    )

    # ---- what the whole plan adds up to ------------------------------------
    first, last = TP_RUNG_FIRST, TP_RUNG_LAST
    planned = f"SUM($H${first}:$H${last})"
    tiles = [
        (2, "Net cash it returns", f"=SUM($K${first}:$K${last})", MONEY0),
        (
            4,
            "Average exit price",
            f'=IF(MIN({planned},{units})=0,"—",'
            f"SUM($J${first}:$J${last})/MIN({planned},{units}))",
            MONEY,
        ),
        (6, "% of position sold", f'=IF({units}=0,"",{planned}/{units})', PCT),
        (8, "Units still riding", f"={units}-{planned}", QTY),
        (
            10,
            "Free ride from",
            f'=IFERROR("+"&TEXT(INDEX($B${first}:$B${last},'
            f'MATCH("Free ride",$N${first}:$N${last},0)),"0%"),"not with this plan")',
            None,
        ),
    ]
    for col, label, formula, fmt in tiles:
        ws.merge_cells(
            start_row=TP_TILE_LABEL,
            start_column=col,
            end_row=TP_TILE_LABEL,
            end_column=col + 1,
        )
        head = ws.cell(row=TP_TILE_LABEL, column=col, value=label.upper())
        head.font = KPI_LABEL_F
        head.fill = PatternFill("solid", fgColor=DERIVED)
        head.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(
            start_row=TP_TILE_VALUE,
            start_column=col,
            end_row=TP_TILE_VALUE + 1,
            end_column=col + 1,
        )
        val = ws.cell(row=TP_TILE_VALUE, column=col, value=formula)
        val.font = Font(name="Calibri", size=15, bold=True, color=INK)
        val.alignment = Alignment(vertical="center", indent=1)
        if fmt:
            val.number_format = fmt
        for offset in range(2):
            for r in (TP_TILE_LABEL, TP_TILE_VALUE, TP_TILE_VALUE + 1):
                ws.cell(row=r, column=col + offset).border = BOX
    ws.row_dimensions[TP_TILE_VALUE].height = 20

    over = PatternFill("solid", fgColor=BAD_SOFT)
    ws.conditional_formatting.add(
        f"F{TP_TILE_VALUE}",
        CellIsRule(
            operator="greaterThan",
            formula=["1"],
            font=Font(size=15, bold=True, color=BAD),
            fill=over,
        ),
    )
    ws.conditional_formatting.add(
        f"H{TP_TILE_VALUE}",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            font=Font(size=15, bold=True, color=BAD),
            fill=over,
        ),
    )

    # ---- the ladder --------------------------------------------------------
    for start, end, text, colour in [
        (2, 5, "The rung", SLATE),
        (6, 7, "If you sold the whole position here", "1D4ED8"),
        (8, 15, "Or take a slice at each rung", ACCENT),
    ]:
        ws.merge_cells(
            start_row=TP_BANNER, start_column=start, end_row=TP_BANNER, end_column=end
        )
        for col in range(start, end + 1):
            ws.cell(row=TP_BANNER, column=col).fill = PatternFill(
                "solid", fgColor=colour
            )
        cell = ws.cell(row=TP_BANNER, column=start, value=text)
        cell.font = HEAD_F
        cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[TP_BANNER].height = 20

    headers = [
        "Gain from your cost",
        "Target price",
        "vs today's price",
        "Profit per unit",
        "Net cash",
        "Net profit",
        "Units to sell",
        "% of remainder",
        "Gross",
        "Net cash",
        "Cash so far",
        "Units left",
        "Status",
        "Break-even of the rest",
    ]
    for offset, label in enumerate(headers):
        cell = ws.cell(row=TP_HEADER, column=2 + offset, value=label)
        cell.font = HEAD_F
        cell.fill = HEAD_FILL
        cell.border = BOX
        cell.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[TP_HEADER].height = 32
    ws.freeze_panes = f"B{TP_RUNG_FIRST}"

    for offset in range(last - first + 1):
        r = first + offset
        seed = TP_RUNGS[offset] if offset < len(TP_RUNGS) else None
        for col in range(2, 16):
            cell = ws.cell(row=r, column=col)
            style_body(cell)
            if offset % 2 == 1:
                cell.fill = BAND_FILL
        for col in (2, 8):
            style_input(ws.cell(row=r, column=col))

        gain = ws.cell(row=r, column=2, value=seed[0] if seed else None)
        gain.number_format = PCT0

        sold_above = "0" if r == first else f"SUM($H${first}:$H{r - 1})"
        remaining = f"MAX(0,{units}-{sold_above})"
        used = f"MIN($H{r},{remaining})"
        qty = ws.cell(
            row=r,
            column=8,
            value=f"=ROUND({seed[1]}*{remaining},6)" if seed else None,
        )
        qty.number_format = QTY

        blank = f'$B{r}=""'
        noqty = f'OR($B{r}="",$H{r}="")'

        ws.cell(
            row=r, column=3, value=f'=IF({blank},"",{cost}*(1+$B{r}))'
        ).number_format = MONEY
        ws.cell(
            row=r, column=4, value=f'=IF(OR({blank},{live}=0),"",$C{r}/{live}-1)'
        ).number_format = PCT
        ws.cell(
            row=r, column=5, value=f'=IF({blank},"",$C{r}-{cost})'
        ).number_format = MONEY
        ws.cell(
            row=r,
            column=6,
            value=(
                f'=IF({blank},"",{units}*$C{r}*(1-FeeRate)'
                f"-MAX(0,($C{r}-{cost})*{units})*TaxRate)"
            ),
        ).number_format = MONEY
        ws.cell(
            row=r, column=7, value=f'=IF({blank},"",$F{r}-{basis})'
        ).number_format = MONEY
        ws.cell(
            row=r,
            column=9,
            value=f'=IF(OR({noqty},{remaining}=0),"",{used}/{remaining})',
        ).number_format = PCT
        ws.cell(
            row=r, column=10, value=f'=IF({noqty},"",{used}*$C{r})'
        ).number_format = MONEY
        ws.cell(
            row=r,
            column=11,
            value=(
                f'=IF({noqty},"",$J{r}*(1-FeeRate)'
                f"-MAX(0,($C{r}-{cost})*{used})*TaxRate)"
            ),
        ).number_format = MONEY
        ws.cell(
            row=r, column=12, value=f'=IF({noqty},"",SUM($K${first}:$K{r}))'
        ).number_format = MONEY
        ws.cell(
            row=r,
            column=13,
            value=f'=IF({noqty},"",MAX(0,{units}-SUM($H${first}:$H{r})))',
        ).number_format = QTY
        ws.cell(
            row=r,
            column=14,
            value=(
                f'=IF({noqty},"",IF(SUM($H${first}:$H{r})>{units},"Oversold",'
                f'IF($L{r}>={basis},"Free ride","Still exposed")))'
            ),
        )
        ws.cell(
            row=r,
            column=15,
            value=f'=IF({noqty},"",IF($M{r}<=0,"—",MAX(0,{basis}-$L{r})/$M{r}))',
        ).number_format = MONEY

    ws.conditional_formatting.add(
        f"N{first}:N{last}",
        CellIsRule(
            operator="equal",
            formula=['"Free ride"'],
            font=Font(color=GOOD, bold=True),
            fill=PatternFill("solid", fgColor=GOOD_SOFT),
        ),
    )
    ws.conditional_formatting.add(
        f"N{first}:N{last}",
        CellIsRule(
            operator="equal",
            formula=['"Oversold"'],
            font=Font(color=BAD, bold=True),
            fill=over,
        ),
    )
    ws.conditional_formatting.add(
        f"D{first}:D{last}",
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["0"],
            font=Font(color=GOOD, bold=True),
            fill=PatternFill("solid", fgColor=GOOD_SOFT),
        ),
    )
    ws.conditional_formatting.add(
        f"L{first}:L{last}",
        DataBarRule(start_type="min", end_type="max", color=ACCENT),
    )

    # ---- recovery math, beside the ladder ----------------------------------
    ws.merge_cells(
        start_row=TP_BANNER, start_column=17, end_row=TP_BANNER, end_column=20
    )
    for col in range(17, 21):
        ws.cell(row=TP_BANNER, column=col).fill = PatternFill("solid", fgColor=BAD)
    rec = ws.cell(row=TP_BANNER, column=17, value="If it goes the other way")
    rec.font = HEAD_F
    rec.alignment = Alignment(vertical="center", horizontal="center")
    for offset, label in enumerate(
        [
            "If you are down",
            "Price becomes",
            "Position worth",
            "Gain needed to get back",
        ]
    ):
        cell = ws.cell(row=TP_HEADER, column=17 + offset, value=label)
        cell.font = HEAD_F
        cell.fill = HEAD_FILL
        cell.border = BOX
        cell.alignment = Alignment(vertical="center", wrap_text=True, indent=1)

    for offset, drawdown in enumerate([i / 100 for i in range(10, 81, 10)]):
        r = TP_RUNG_FIRST + offset
        for col in range(17, 21):
            cell = ws.cell(row=r, column=col)
            style_body(cell)
            if offset % 2 == 1:
                cell.fill = BAND_FILL
        ws.cell(row=r, column=17, value=round(drawdown, 2)).number_format = PCT0
        ws.cell(row=r, column=18, value=f"={cost}*(1-$Q{r})").number_format = MONEY
        ws.cell(row=r, column=19, value=f"=$R{r}*{units}").number_format = MONEY
        ws.cell(row=r, column=20, value=f"=1/(1-$Q{r})-1").number_format = PCT0
    ws.conditional_formatting.add(
        f"T{TP_RUNG_FIRST}:T{TP_RUNG_FIRST + 7}",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["1"],
            font=Font(color=BAD, bold=True),
            fill=over,
        ),
    )

    # The rung the market is heading for next, resolved here rather than in the
    # watcher so the sheet and the alert can never disagree about it.
    nxt = TP_RUNG_FIRST + 9
    ws.cell(row=nxt, column=17, value="NEXT RUNG").font = KPI_LABEL_F
    next_block = [
        (
            "which",
            # Capped at the number of rungs that exist, not the number of rows:
            # a price past every rung used to land on a blank one and report
            # no target at all.
            f"=IFERROR(MIN(MATCH(0,$D${first}:$D${last},1)+1,"
            f"COUNT($C${first}:$C${last})),1)",
            "0",
        ),
        ("gain", f"=INDEX($B${first}:$B${last},$R{nxt})", PCT0),
        ("target price", f"=INDEX($C${first}:$C${last},$R{nxt})", MONEY),
        ("away", f"=INDEX($D${first}:$D${last},$R{nxt})", PCT),
        ("units to sell", f"=INDEX($H${first}:$H${last},$R{nxt})", QTY),
        ("net cash", f"=INDEX($K${first}:$K${last},$R{nxt})", MONEY),
        ("units left after", f"=INDEX($M${first}:$M${last},$R{nxt})", QTY),
        ("break-even after", f"=INDEX($O${first}:$O${last},$R{nxt})", MONEY),
    ]
    for offset, (label, formula, fmt) in enumerate(next_block):
        r = nxt + offset
        lab = ws.cell(row=r, column=17, value=label if offset else "which rung")
        lab.font = SMALL_F
        cell = ws.cell(row=r, column=18, value=formula)
        style_derived(cell)
        cell.number_format = fmt

    note = ws.cell(
        row=last + 2,
        column=2,
        value=(
            "The first amber cell lists the names on your Positions tab. Pick "
            "one and this plan reads its ticker, quantity and average cost "
            "from that row — it is not a coin type and not a ticker symbol, "
            "it is the name of a row you keep. Type over what it pulls in for "
            "a what-if and the tab says EDITED; name something that is not on "
            "Positions and it says so in red.\n"
            "Every unit cell is a share of what the rungs above it left, so "
            "changing one re-sizes the rest against what actually remains. "
            "Price mode on Manual pins Your price over the feed."
        ),
    )
    note.font = SMALL_F
    note.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=last + 2, start_column=2, end_row=last + 3, end_column=11)
    return ws


def build_trade_log(wb: Workbook):
    ws = wb.create_sheet("Trade Log")
    ws.sheet_properties.tabColor = "1D4ED8"
    sheet_title(
        ws,
        "Trade Log",
        "Every fill you actually made. Pick the position, give it a round "
        "number, and each sell works out what it realised against what that "
        "round's buys averaged.",
        17,
    )
    headers = [
        "Date",
        "Position",
        "Round",
        "Side",
        "Units",
        "Price",
        "Fees",
        "Cash flow",
        "Buy cost",
        "Buy units",
        "Round avg cost",
        "Realised $",
        "Realised %",
        "Thesis at entry",
        "What would prove me wrong",
        "Exit reason",
        "Followed the plan?",
    ]
    widths = [12, 26, 8, 9, 14, 13, 10, 14, 12, 11, 15, 13, 12, 40, 36, 28, 15]
    header_row(ws, 4, headers, widths)
    ws.freeze_panes = "C5"

    asset_dv = DataValidation(
        type="list",
        formula1=f"=Positions!$A${POS_FIRST}:$A${POS_LAST}",
        allow_blank=True,
    )
    side_dv = DataValidation(type="list", formula1="=Lists!$C$2:$C$3", allow_blank=True)
    yesno_dv = DataValidation(
        type="list", formula1="=Lists!$D$2:$D$3", allow_blank=True
    )
    for dv in (asset_dv, side_dv, yesno_dv):
        ws.add_data_validation(dv)

    for row in range(LOG_FIRST, LOG_LAST + 1):
        for col in range(1, 18):
            cell = ws.cell(row=row, column=col)
            style_body(cell)
            if (row - LOG_FIRST) % 2 == 1:
                cell.fill = BAND_FILL
        for col in (1, 2, 3, 4, 5, 6, 7, 14, 15, 16, 17):
            style_input(ws.cell(row=row, column=col))
        for col in (9, 10, 11):
            style_derived(ws.cell(row=row, column=col))

        ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5).number_format = QTY
        for col in (6, 7):
            ws.cell(row=row, column=col).number_format = MONEY
        asset_dv.add(ws.cell(row=row, column=2))
        side_dv.add(ws.cell(row=row, column=4))
        yesno_dv.add(ws.cell(row=row, column=17))

        empty = f'OR($E{row}="",$F{row}="")'
        ws.cell(
            row=row,
            column=8,
            value=(
                f'=IF({empty},"",IF($D{row}="Buy",-($E{row}*$F{row})-N($G{row}),'
                f"($E{row}*$F{row})-N($G{row})))"
            ),
        ).number_format = MONEY
        # These two exist so the round average below can be a SUMIFS rather
        # than an array formula. Nothing else reads them.
        ws.cell(
            row=row,
            column=9,
            value=f'=IF(OR($D{row}<>"Buy",{empty}),"",$E{row}*$F{row}+N($G{row}))',
        ).number_format = MONEY
        ws.cell(
            row=row,
            column=10,
            value=f'=IF(OR($D{row}<>"Buy",{empty}),"",$E{row})',
        ).number_format = QTY
        # Average cost of this position's buys in this round only, which is
        # what makes re-entering a closed position safe: round 2 never sees
        # round 1's prices.
        ws.cell(
            row=row,
            column=11,
            value=(
                f'=IF($B{row}="","",IFERROR('
                f"SUMIFS($I${LOG_FIRST}:$I${LOG_LAST},$B${LOG_FIRST}:$B${LOG_LAST},$B{row},"
                f"$C${LOG_FIRST}:$C${LOG_LAST},$C{row})/"
                f"SUMIFS($J${LOG_FIRST}:$J${LOG_LAST},$B${LOG_FIRST}:$B${LOG_LAST},$B{row},"
                f'$C${LOG_FIRST}:$C${LOG_LAST},$C{row}),""))'
            ),
        ).number_format = MONEY
        ws.cell(
            row=row,
            column=12,
            value=(
                f'=IF(OR($D{row}<>"Sell",{empty},$K{row}=""),"",'
                f"$E{row}*$F{row}-N($G{row})-$E{row}*$K{row})"
            ),
        ).number_format = MONEY
        ws.cell(
            row=row,
            column=13,
            value=f'=IF(OR($L{row}="",$K{row}=0),"",$L{row}/($E{row}*$K{row}))',
        ).number_format = PCT
        for col in (14, 15, 16):
            ws.cell(row=row, column=col).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    for span in (f"L{LOG_FIRST}:M{LOG_LAST}",):
        ws.conditional_formatting.add(
            span,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                font=Font(color=GOOD, bold=True),
                fill=PatternFill("solid", fgColor=GOOD_SOFT),
            ),
        )
        ws.conditional_formatting.add(
            span,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                font=Font(color=BAD, bold=True),
                fill=PatternFill("solid", fgColor=BAD_SOFT),
            ),
        )
    ws.conditional_formatting.add(
        f"Q{LOG_FIRST}:Q{LOG_LAST}",
        CellIsRule(
            operator="equal",
            formula=['"No"'],
            font=Font(color=BAD, bold=True),
            fill=PatternFill("solid", fgColor=BAD_SOFT),
        ),
    )
    return ws


def build_watch(wb: Workbook):
    """One flat table for the alert watcher to read.

    Published to the web as CSV and polled from the PC. It exists so the
    watcher reads a single tab instead of seven, and so the sheet — not the
    script — decides which rung is next: two places computing the same target
    is how they end up disagreeing.
    """
    ws = wb.create_sheet("Watch")
    ws.sheet_properties.tabColor = MUTED
    sheet_title(
        ws,
        "Watch",
        "What the alert watcher reads. Publish this one tab to the web as CSV. "
        "Nothing here is typed.",
        13,
    )

    headers = [
        "Plan",
        "Holding",
        "Ticker",
        "Feed",
        "Units held",
        "Avg cost",
        "Price",
        "Weight",
        "Next rung",
        "Target price",
        "Away",
        "Units to sell",
        "Net cash",
    ]
    header_row(ws, 3, headers, [10, 26, 20, 22, 14, 13, 13, 10, 11, 14, 10, 14, 13])

    nxt = TP_RUNG_FIRST + 9
    for index in range(1, PLAN_COUNT + 1):
        row = 3 + index
        plan = f"'Plan {index}'"
        holding = f"{plan}!$B${TP_INPUT}"
        for col in range(1, 14):
            style_derived(ws.cell(row=row, column=col))
        cells = [
            (1, f'="Plan {index}"', None),
            (2, f'=IF({holding}="","",{holding})', None),
            (3, f'=IF({holding}="","",{plan}!$C${TP_INPUT})', None),
            (4, f'=IF({holding}="","",{plan}!$I${TP_INPUT})', None),
            (5, f'=IF({holding}="","",{plan}!$D${TP_INPUT})', QTY),
            (6, f'=IF({holding}="","",{plan}!$E${TP_INPUT})', MONEY),
            (7, f'=IF({holding}="","",{plan}!$H${TP_INPUT})', MONEY),
            (
                8,
                f'=IF({holding}="","",IFERROR(INDEX(Positions!$Q${POS_FIRST}:$Q${POS_LAST},'
                f"MATCH({holding},Positions!$A${POS_FIRST}:$A${POS_LAST},0)),"
                "))",
                PCT,
            ),
            (9, f'=IF({holding}="","",{plan}!$R${nxt + 1})', PCT0),
            (10, f'=IF({holding}="","",{plan}!$R${nxt + 2})', MONEY),
            (11, f'=IF({holding}="","",{plan}!$R${nxt + 3})', PCT),
            (12, f'=IF({holding}="","",{plan}!$R${nxt + 4})', QTY),
            (13, f'=IF({holding}="","",{plan}!$R${nxt + 5})', MONEY),
        ]
        for col, formula, fmt in cells:
            cell = ws.cell(row=row, column=col, value=formula)
            if fmt:
                cell.number_format = fmt

    row = 3 + PLAN_COUNT + 2
    ws.cell(row=row, column=1, value="Holdings").font = BOLD_F
    header_row_at = row + 1
    for offset, (label, width) in enumerate(
        [
            ("Holding", 26),
            ("Ticker", 20),
            ("Class", 12),
            ("Status", 10),
            ("Feed", 14),
            ("Units", 14),
            ("Avg cost", 13),
            ("Price", 13),
            ("Market value", 14),
            ("Weight", 10),
            ("Unrealised %", 13),
        ]
    ):
        cell = ws.cell(row=header_row_at, column=1 + offset, value=label)
        cell.font = HEAD_F
        cell.fill = HEAD_FILL
        cell.border = BOX
        cell.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[header_row_at].height = 28

    for offset in range(POS_LAST - POS_FIRST + 1):
        src = POS_FIRST + offset
        row = header_row_at + 1 + offset
        for col, letter, fmt in (
            (1, "A", None),
            (2, "B", None),
            (3, "C", None),
            (4, "E", None),
            (5, "M", None),
            (6, "G", QTY),
            (7, "H", MONEY),
            (8, "L", MONEY),
            (9, "N", MONEY),
            (10, "Q", PCT),
            (11, "P", PCT),
        ):
            cell = ws.cell(
                row=row,
                column=col,
                value=f'=IF(Positions!$A${src}="","",Positions!${letter}${src})',
            )
            style_derived(cell)
            if fmt:
                cell.number_format = fmt
    return ws


def build_dashboard(wb: Workbook):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = INK
    ws.sheet_view.showGridLines = False
    sheet_title(
        ws, "Dashboard", "Everything here is calculated. Nothing here is typed.", 18
    )
    ws.cell(
        row=2,
        column=1,
        value='=Settings!$C$7&" — everything here is calculated, nothing here is typed."',
    ).font = SUB_F

    ws.column_dimensions["A"].width = 3
    for col in "BCDFGHJKLNOP":
        ws.column_dimensions[col].width = 12
    for col in "EIM":
        ws.column_dimensions[col].width = 2

    status = f"Positions!$E${POS_FIRST}:$E${POS_LAST}"
    realised = f"Positions!$R${POS_FIRST}:$R${POS_LAST}"
    # Value sitting on prices nobody is updating. Counting the rows understates
    # it badly — one stale holding was a third of the book while reading as one
    # row out of twelve.
    stale = (
        f'SUMIF(Positions!$M${POS_FIRST}:$M${POS_LAST},"<>live",'
        f"Positions!$N${POS_FIRST}:$N${POS_LAST})"
    )

    def tiles(label_row: int, entries):
        for start_col, label, formula, fmt, size in entries:
            col = ws[f"{start_col}1"].column
            ws.merge_cells(
                start_row=label_row,
                start_column=col,
                end_row=label_row,
                end_column=col + 2,
            )
            head = ws.cell(row=label_row, column=col, value=label.upper())
            head.font = KPI_LABEL_F
            head.alignment = Alignment(vertical="center", indent=1)
            head.fill = PatternFill("solid", fgColor=DERIVED)
            ws.merge_cells(
                start_row=label_row + 1,
                start_column=col,
                end_row=label_row + 2,
                end_column=col + 2,
            )
            val = ws.cell(row=label_row + 1, column=col, value=formula)
            val.font = Font(name="Calibri", size=size, bold=True, color=INK)
            val.number_format = fmt or "General"
            val.alignment = Alignment(vertical="center", indent=1)
            for offset in range(3):
                for r in (label_row, label_row + 1, label_row + 2):
                    ws.cell(row=r, column=col + offset).border = BOX
        ws.row_dimensions[label_row + 1].height = 22
        ws.row_dimensions[label_row + 2].height = 16

    live_value = (
        f'SUMIF(Positions!$M${POS_FIRST}:$M${POS_LAST},"live",'
        f"Positions!$N${POS_FIRST}:$N${POS_LAST})"
    )
    tiles(
        4,
        [
            # Split, so the headline is never contaminated. A number nobody is
            # updating does not belong in the same total as a live one.
            ("B", "Value on live prices", f"={live_value}", MONEY0, 20),
            ("F", "Value priced by hand", f"={stale}", MONEY0, 20),
            ("J", "Unrealised", f"=Positions!$O${POS_TOTAL}", MONEY0, 20),
            (
                "N",
                "Unrealised %",
                f"=IFERROR(Positions!$O${POS_TOTAL}/Positions!$I${POS_TOTAL},0)",
                PCT,
                20,
            ),
        ],
    )
    # Directly under the portfolio value, because that is the number the
    # staleness makes untrue.
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=16)
    band = ws.cell(
        row=7,
        column=2,
        value=(
            f'=IF(Positions!$N${POS_TOTAL}=0,"Nothing in the register yet — '
            f'add a holding on Positions and every tab fills in.",'
            f'IF({stale}=0,"Every position is priced from a live feed.",'
            f'"Portfolio "&TEXT(Positions!$N${POS_TOTAL},"$#,##0")&", of which "'
            f'&TEXT({stale},"$#,##0")&" — "'
            f'&TEXT(IFERROR({stale}/Positions!$N${POS_TOTAL},0),"0.0%")'
            f'&" — is priced by hand and has not moved since you typed it."))'
        ),
    )
    band.font = Font(name="Calibri", size=10, bold=True, color=INK)
    band.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[7].height = 20
    ws.conditional_formatting.add(
        "B7",
        FormulaRule(
            formula=[f"{stale}>0"],
            fill=PatternFill("solid", fgColor=WARN_SOFT),
            font=Font(size=10, bold=True, color="92400E"),
        ),
    )
    ws.conditional_formatting.add(
        "B7",
        FormulaRule(
            formula=[f"{stale}=0"],
            fill=PatternFill("solid", fgColor=GOOD_SOFT),
            font=Font(size=10, bold=True, color=GOOD),
        ),
    )

    # The second row is the half the old sheet never had: what you have
    # actually banked, and what the two halves add up to.
    tiles(
        8,
        [
            ("B", "Realised, all time", f"=Positions!$R${POS_TOTAL}", MONEY0, 18),
            ("F", "Total result", f"=Positions!$S${POS_TOTAL}", MONEY0, 18),
            (
                "J",
                "Active / closed",
                f'=COUNTIF({status},"Active")&" / "&COUNTIF({status},"Closed")',
                None,
                18,
            ),
            (
                "N",
                "Closed rounds in profit",
                f'=IFERROR(COUNTIFS({status},"Closed",{realised},">0")'
                f'/COUNTIF({status},"Closed"),"—")',
                PCT0,
                18,
            ),
        ],
    )
    for span in ("J5:P7", "B9:H11"):
        ws.conditional_formatting.add(
            span,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                font=Font(size=18, bold=True, color=BAD),
            ),
        )
        ws.conditional_formatting.add(
            span,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                font=Font(size=18, bold=True, color=GOOD),
            ),
        )

    def breakdown(col: int, heading: str, keys, key_range: str):
        title = ws.cell(row=12, column=col, value=heading)
        title.font = BOLD_F
        for offset, (label, width) in enumerate(
            [("", 22), ("Value", 14), ("Share", 10)]
        ):
            cell = ws.cell(
                row=13, column=col + offset, value=label or heading.split()[-1]
            )
            cell.font = HEAD_F
            cell.fill = HEAD_FILL
            cell.border = BOX
            cell.alignment = Alignment(vertical="center", indent=1)
            ws.column_dimensions[get_column_letter(col + offset)].width = width
        letter = get_column_letter(col)
        value_letter = get_column_letter(col + 1)
        for offset, key in enumerate(keys):
            row = 14 + offset
            for c in range(col, col + 3):
                cell = ws.cell(row=row, column=c)
                style_body(cell)
                if offset % 2 == 1:
                    cell.fill = BAND_FILL
            ws.cell(row=row, column=col, value=key)
            ws.cell(
                row=row,
                column=col + 1,
                value=(
                    f"=SUMIF({key_range},${letter}{row},"
                    f"Positions!$N${POS_FIRST}:$N${POS_LAST})"
                ),
            ).number_format = MONEY0
            ws.cell(
                row=row,
                column=col + 2,
                value=f'=IFERROR(${value_letter}{row}/Positions!$N${POS_TOTAL},"")',
            ).number_format = PCT
        last = 14 + len(keys) - 1
        ws.conditional_formatting.add(
            f"{get_column_letter(col + 2)}14:{get_column_letter(col + 2)}{last}",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=1,
                color=ACCENT,
            ),
        )
        return last

    class_last = breakdown(
        2,
        "Split by class",
        CLASSES,
        f"Positions!$C${POS_FIRST}:$C${POS_LAST}",
    )
    sector_last = breakdown(
        6,
        "Split by sector",
        SECTORS,
        f"Positions!$D${POS_FIRST}:$D${POS_LAST}",
    )

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Where the money actually is"
    chart.height = 10
    chart.width = 16
    chart.add_data(
        Reference(ws, min_col=7, min_row=13, max_row=sector_last), titles_from_data=True
    )
    chart.set_categories(Reference(ws, min_col=6, min_row=14, max_row=sector_last))
    chart.legend = None
    ws.add_chart(chart, "J13")

    row = max(class_last, sector_last) + 3
    ws.cell(row=row, column=2, value="Checks").font = BOLD_F
    checks = [
        (
            "Active holdings with no plan",
            f'=SUMPRODUCT(({status}="Active")*'
            f"(COUNTIF(Watch!$B$4:$B${3 + PLAN_COUNT},"
            f"Positions!$A${POS_FIRST}:$A${POS_LAST})=0))",
            f"Every holding should have rungs before it needs them. There are "
            f"{PLAN_COUNT} plan tabs; duplicate one for a seventh coin.",
        ),
        (
            "Rows the register is ignoring",
            f'=COUNTIF(Positions!$Y${POS_FIRST}:$Y${POS_LAST},"set*")',
            "A holding missing a Status, Quantity, Avg cost, Class or Ticker. It counts toward nothing until you finish it, and the row says which field it wants.",
        ),
        (
            "Plans that no longer match the register",
            "+".join(
                f"IF('Plan {i}'!$N${TP_INPUT}=\"EDITED\",1,0)"
                for i in range(1, PLAN_COUNT + 1)
            ).join(("=", "")),
            "You typed over a quantity or cost the plan pulled from Positions. Fine for a what-if, misleading if you forgot.",
        ),
        (
            "Portfolio value priced by hand",
            f"={stale}",
            "Not a row count — the money riding on prices that do not update themselves. The band under the tiles says what share of the portfolio that is.",
        ),
        (
            "Trades logged against nothing",
            f"=SUMPRODUCT(('Trade Log'!$B${LOG_FIRST}:$B${LOG_LAST}<>\"\")*"
            f"(COUNTIF(Positions!$A${POS_FIRST}:$A${POS_LAST},"
            f"'Trade Log'!$B${LOG_FIRST}:$B${LOG_LAST})=0))",
            "A logged trade whose position name does not match the register earns nothing — its realised profit lands nowhere.",
        ),
        (
            "Names used twice on the register",
            f'=SUMPRODUCT((Positions!$A${POS_FIRST}:$A${POS_LAST}<>"")*'
            f"(COUNTIF(Positions!$A${POS_FIRST}:$A${POS_LAST},"
            f"Positions!$A${POS_FIRST}:$A${POS_LAST})>1))",
            "Names are the key everything else joins on. Re-entering a position needs a name you can tell apart — Cardano, round 2 — not a second row called Cardano.",
        ),
        (
            "Trades logged without a thesis",
            f"=SUMPRODUCT(('Trade Log'!$B${LOG_FIRST}:$B${LOG_LAST}<>\"\")"
            f"*('Trade Log'!$N${LOG_FIRST}:$N${LOG_LAST}=\"\"))",
            "A trade with no written reason cannot be reviewed later. It is just a number.",
        ),
    ]
    for offset, (label, formula, note) in enumerate(checks, start=1):
        r = row + offset
        ws.cell(row=r, column=2, value=label).font = BODY_F
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        val = ws.cell(row=r, column=4, value=formula)
        val.font = BOLD_F
        val.number_format = MONEY0 if "SUMIF(Positions!$M" in formula else "0"
        val.alignment = Alignment(horizontal="center")
        val.border = BOX
        hint = ws.cell(row=r, column=6, value=note)
        hint.font = SMALL_F
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=14)
        ws.conditional_formatting.add(
            f"D{r}",
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                font=Font(color=BAD, bold=True),
                fill=PatternFill("solid", fgColor=WARN_SOFT),
            ),
        )
        ws.conditional_formatting.add(
            f"D{r}",
            CellIsRule(
                operator="equal",
                formula=["0"],
                font=Font(color=GOOD, bold=True),
                fill=PatternFill("solid", fgColor=GOOD_SOFT),
            ),
        )
    return ws


def build_start_here(wb: Workbook):
    ws = wb.create_sheet("Start Here")
    ws.sheet_properties.tabColor = ACCENT
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 104

    ws.merge_cells("B1:C1")
    title = ws.cell(row=1, column=2, value="Profit & Exit Planner")
    title.font = Font(name="Calibri", size=22, bold=True, color=WHITE)
    title.fill = TITLE_FILL
    title.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 44

    def section(row: int, heading: str) -> int:
        cell = ws.cell(row=row, column=2, value=heading)
        cell.font = Font(name="Calibri", size=12, bold=True, color=ACCENT)
        return row + 1

    def line(row: int, left: str, right: str) -> int:
        lab = ws.cell(row=row, column=2, value=left)
        lab.font = BOLD_F
        lab.alignment = Alignment(vertical="top")
        body = ws.cell(row=row, column=3, value=right)
        body.font = BODY_F
        body.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = max(16, 14 * (1 + len(right) // 95))
        return row + 1

    row = 3
    row = section(row, "What this is")
    row = line(
        row,
        "",
        "A rebuild of the old 'Profit Gain Percentage' sheet. Same job — decide "
        "what to sell, at what price, and what it leaves you with — but the "
        "numbers are calculated instead of typed, so they stop being wrong the "
        "moment the market moves.",
    )
    row += 1

    row = section(row, "Start on a Plan tab")
    row = line(
        row,
        "Plan 1 to Plan 6",
        "Pick a holding from the dropdown and the register fills in the "
        "ticker, quantity and average cost. Each rung then shows what the "
        "price has to reach, what you make per unit, what selling the lot "
        "would return, and what taking a slice leaves you holding. Six ship "
        "ready-made; duplicate one for a seventh coin.",
    )
    row += 1

    row = section(row, "The rest")
    row = line(
        row,
        "Positions",
        "The register. Every holding lives here once, and the plans read from "
        "it — which is what stops the same coin existing at two different "
        "average costs. A row missing something says so in its last column and "
        "turns amber.",
    )
    row = line(
        row,
        "Dashboard",
        "Totals and the checks that go red: holdings with no plan, rows the "
        "register is ignoring, plans that no longer match it.",
    )
    row = line(
        row,
        "Trade Log",
        "One row per fill, with the thesis and the invalidation written down at the time.",
    )
    row = line(
        row,
        "Watch",
        "What the alert watcher reads. Nothing to fill in — publish this one "
        "tab to the web as CSV and the watcher on your PC polls it.",
    )
    row = line(
        row,
        "Settings",
        "Tax rate, fees, weight limit. Four numbers the other tabs read.",
    )
    row += 1

    row = section(row, "How prices work")
    row = line(
        row,
        "",
        "Each position calls GOOGLEFINANCE on its ticker. When the feed does not "
        "carry a symbol — it does not carry every altcoin — the price falls back "
        "to the manual price beside it and the Feed column says so. Set Price "
        "mode to Manual to pin a price deliberately. The Dashboard counts how "
        "many rows are running on hand-typed prices.",
    )
    row += 1

    row = section(row, "Colour code")
    for bg, edge, text in (
        (INPUT_BG, INPUT_LINE, "Amber — type here."),
        (
            DERIVED,
            LINE,
            "Grey italic — pulled from another tab. Overwrite it and the link breaks.",
        ),
        (WHITE, LINE, "White — calculated. Leave it alone and it stays right."),
    ):
        chip = ws.cell(row=row, column=2, value="")
        chip.fill = PatternFill("solid", fgColor=bg)
        side = Side(style="thin", color=edge)
        chip.border = Border(left=side, right=side, top=side, bottom=side)
        desc = ws.cell(row=row, column=3, value=text)
        desc.font = BODY_F
        row += 1
    row += 1

    row = section(row, "What changed from the old sheet")
    changes = [
        (
            "Duplicates removed",
            "Decentraland and Chainlink were listed twice, Cardano's exit plan three times, and two 'Future Asset' placeholder rows sat among the real ones. Sixteen rows became twelve holdings.",
        ),
        (
            "A ladder that oversold",
            "The Gala plan sold 3,563 coins on its last rung with 891 left, booking roughly $374 of profit on coins already gone. Each rung is now a share of what the rungs above it left, so a plan cannot sell what you no longer hold.",
        ),
        (
            "The percentage grid",
            "About 1,900 hand-typed cells across 14 blocks, all of them the same arithmetic. It is now a handful of cells on a Plan tab, and the rungs move when you change what you paid.",
        ),
        (
            "Six tabs became one",
            "Working out an exit used to mean reading a percentage table, a sell-profit table and a pull-profits block that did not agree with each other. A Plan tab answers all three questions in one row: what the price has to reach, what selling the lot returns, and what selling a slice leaves you holding.",
        ),
        (
            "Prices unfroze",
            "Every price was a value typed in October 2022. They are formulas now.",
        ),
        (
            "An all-time high of $0.00",
            "Shiba Inu's real high, $0.00008845, had been rounded away by the cell format, which made its X-to-ATH meaningless.",
        ),
        (
            "Cost basis is one number",
            "The old sheet stored price and cost separately and they had drifted apart. Average cost is stored once and cost basis is derived.",
        ),
        (
            "Tax and fees exist",
            "Every exit figure is now shown after both. The old sheet's profits were gross.",
        ),
        (
            "Recovery math added",
            "Being down 50% needs +100% to undo. The old sheet only ever counted upward.",
        ),
        (
            "Break-even of the remainder",
            "After selling a slice, what the rest has to be worth for the trade to be flat — and when it drops through zero, the position can no longer lose you money.",
        ),
        (
            "The trading log",
            "Was 121 dated rows with four side-by-side ticker blocks and nothing filled in. Now one row per fill, with the thesis and the invalidation captured at entry, and each sell measured against what that round's buys averaged.",
        ),
        (
            "Profit you have actually banked",
            "The old sheet only ever showed what a position might be worth. Closing one out now leaves its realised result on the register, out of the portfolio totals but inside the lifetime figure — so selling something does not erase it from your record.",
        ),
        (
            "Credentials removed",
            "The old sheet had a recovery key and a password sitting in plain cells. They were not copied here. Treat them as exposed and rotate them.",
        ),
        (
            "Not carried over",
            "Avalanche, Polygon and Cronos appeared only on planning tabs, never on the holdings tab. They are not here as positions — add them if you hold them.",
        ),
    ]
    for label, text in changes:
        row = line(row, label, text)
    row += 1

    row = section(row, "The disclaimer, kept from the old sheet")
    row = line(
        row,
        "",
        "This spreadsheet is for informational and educational purposes ONLY. I "
        "am NOT a financial advisor and this is NOT financial advice :-)",
    )
    row += 1
    footer = ws.cell(
        row=row,
        column=3,
        value="Generated by tools/build_profit_planner.py in pwb-toolbox. Rebuild it rather than patching it.",
    )
    footer.font = SMALL_F
    return ws


def build_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    # Created in the order they should appear: the daily driver first, the
    # reference tabs behind it.
    build_start_here(wb)
    for index in range(1, PLAN_COUNT + 1):
        build_plan_tab(wb, index)
    build_positions(wb)
    build_dashboard(wb)
    build_trade_log(wb)
    build_watch(wb)
    build_settings(wb)
    build_lists(wb)
    wb.active = 0
    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--out",
        default="profit_planner.xlsx",
        help="path to write the workbook to (default: profit_planner.xlsx)",
    )
    args = parser.parse_args()
    build_workbook().save(args.out)
    print(f"wrote {args.out}")


if __name__ == "__main__":
    main()
