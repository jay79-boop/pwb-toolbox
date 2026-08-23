#!/usr/bin/env python
"""Spicy Lab's two supports: the Excel ladder export and the quote helper.

The interactive instrument is `static/spicy-lab.html` (opens from file://,
math from `static/option-lab.js`). This tool covers what a browser page
cannot do by itself:

    excel     Write the move ladder for one contract to a spreadsheet — the
              planner-style artifact: rungs of underlying moves as rows, time
              marching right as columns, each cell the position P&L, plus a
              summary block with the greeks, shot clock, and hourly hurdle.

    serve     The local quote helper. A stdlib HTTP server on
              127.0.0.1:8877 that answers /quote?symbol=SPY with the latest
              price via yfinance, with CORS headers so the file:// lab page
              can call it. The page's Refresh button lights up only while
              this is running; without it the lab is type-in, which also
              works. Loopback only — never exposed off the machine.

All pricing goes through `pwb_toolbox.options` — the same tested module the
JS lab mirrors — so the spreadsheet and the page cannot disagree.

Examples::

    python tools/spicy_lab.py excel --spot 640 --strike 640 --days 1.27 --iv 16 --out spy_ladder.xlsx
    python tools/spicy_lab.py excel --spot 640 --strike 645 --days 0.27 --iv 18 --kind call --premium 1.10 --qty 2 --out zero_dte.xlsx
    python tools/spicy_lab.py serve
"""

from __future__ import annotations

import argparse
import json
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import parse_qs, urlparse

from pwb_toolbox.options import black_scholes, expected_move, shot_clock

MINUTES_COLS = [0, 15, 30, 60, 120]
EM_RUNGS = [2, 1.5, 1, 0.5, 0.25, 0, -0.25, -0.5, -1, -1.5, -2]
PCT_RUNGS = [3, 2, 1, 0.5, 0.25, 0, -0.25, -0.5, -1, -2, -3]
HELPER_PORT = 8877


def trading_hours_left(days: float) -> float:
    """6.5 trading hours per whole day plus the partial day's clock hours,
    capped at one session. Approximate on purpose; the same rule as the lab."""
    whole = int(days)
    return whole * 6.5 + min((days - whole) * 24, 6.5)


def ladder_rows(
    spot, strike, days, iv, kind="call", premium=None, contracts=1, mode="em"
):
    """The move ladder as plain rows: one per rung, with position P&L at each
    minutes-elapsed column. `premium` defaults to the model price so the
    ladder works before a fill exists."""
    entry = black_scholes(spot, strike, days, iv, kind=kind)
    if premium is None:
        premium = entry.price
    em = expected_move(spot, iv, min(days, 1.0))
    em_pct = em / spot * 100
    rungs = (
        [(k, k * em_pct) for k in EM_RUNGS]
        if mode == "em"
        else [(k, k) for k in PCT_RUNGS]
    )
    cost = premium * contracts * 100
    rows = []
    for k, pct in rungs:
        level = spot * (1 + pct / 100)
        cells = []
        for minutes in MINUTES_COLS:
            d = days - minutes / 1440
            if d > 0:
                prem = black_scholes(level, strike, d, iv, kind=kind).price
            else:
                prem = max(0.0, strike - level if kind == "put" else level - strike)
            pnl = (prem - premium) * contracts * 100
            cells.append(
                {
                    "minutes": minutes,
                    "premium": round(prem, 4),
                    "pnl": round(pnl, 2),
                    "pnl_pct": round(pnl / cost * 100, 1) if cost > 0 else 0.0,
                }
            )
        rows.append(
            {
                "rung": (
                    "entry"
                    if k == 0
                    else (f"{k:+g}xEM" if mode == "em" else f"{k:+g}%")
                ),
                "move_pct": round(pct, 3),
                "level": round(level, 2),
                "cells": cells,
                "delta_now": round(
                    black_scholes(level, strike, days, iv, kind=kind).delta, 3
                ),
            }
        )
    return rows


def summary_block(
    spot, strike, days, iv, kind="call", premium=None, contracts=1, budget=0.10
):
    g = black_scholes(spot, strike, days, iv, kind=kind)
    if premium is None:
        premium = g.price
    hours = trading_hours_left(days)
    return {
        "premium": round(premium, 4),
        "cost": round(premium * contracts * 100, 2),
        "delta": round(g.delta, 3),
        "gamma": round(g.gamma, 5),
        "theta_day": round(g.theta * contracts * 100, 2),
        "vega_pt": round(g.vega * contracts * 100, 2),
        "shot_clock_h": round(shot_clock(hours, budget), 2),
        "hourly_hurdle": round(expected_move(spot, iv, 1 / 6.5), 2),
        "expected_move": round(expected_move(spot, iv, min(days, 1.0)), 2),
    }


def cmd_excel(args):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    iv = args.iv / 100
    kw = dict(
        spot=args.spot,
        strike=args.strike,
        days=args.days,
        iv=iv,
        kind=args.kind,
        premium=args.premium,
        contracts=args.qty,
    )
    summary = summary_block(**kw, budget=args.budget / 100)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ladder"
    bold = Font(bold=True)

    ws.append(
        [
            f"{args.symbol} {args.strike} {args.kind.upper()}  ({args.days}d, IV {args.iv}%)"
        ]
    )
    ws["A1"].font = bold
    ws.append([])
    for label, key, note in [
        ("Premium / contract", "premium", "entry (model if no fill given)"),
        ("Position cost = max loss", "cost", f"{args.qty} contract(s)"),
        ("Delta", "delta", "$ per $1 of underlying, per share"),
        ("Theta / day", "theta_day", "position $ lost per flat day"),
        ("Vega / IV pt", "vega_pt", "position $ per IV point"),
        (
            "Shot clock (hours)",
            "shot_clock_h",
            f"flat time to lose {args.budget:g}% of premium",
        ),
        ("Hourly hurdle ($)", "hourly_hurdle", "move/hour the underlying owes you"),
        ("Expected move ($)", "expected_move", "1-sigma over the hold window"),
    ]:
        ws.append([label, summary[key], note])
        ws.cell(ws.max_row, 1).font = bold
    ws.append([])

    for mode, title in [("em", "Rungs x expected move"), ("pct", "Fixed % rungs")]:
        rows = ladder_rows(mode=mode, **kw)
        ws.append([title])
        ws.cell(ws.max_row, 1).font = bold
        header = (
            ["rung", "level"]
            + [("now" if m == 0 else f"+{m}m") for m in MINUTES_COLS]
            + ["delta now"]
        )
        ws.append(header)
        for c in range(1, len(header) + 1):
            ws.cell(ws.max_row, c).font = bold
        for r in rows:
            ws.append(
                [r["rung"], r["level"]]
                + [c["pnl"] for c in r["cells"]]
                + [r["delta_now"]]
            )
        ws.append([])

    wb.save(args.out)
    print(f"Ladder -> {args.out}")
    print(
        f"Shot clock {summary['shot_clock_h']}h at a {args.budget:g}% decay budget; "
        f"hourly hurdle ${summary['hourly_hurdle']}."
    )


# ---------------------------------------------------------------------------
# Quote helper
# ---------------------------------------------------------------------------


def quote_payload(symbol: str, fetch=None) -> dict:
    """The /quote response as a dict. `fetch(symbol) -> float` is injectable
    so the handler logic is testable without a market."""
    symbol = (symbol or "").strip().upper()
    if not symbol:
        return {"error": "no symbol"}
    if fetch is None:
        fetch = _yfinance_last
    try:
        price = float(fetch(symbol))
    except Exception as exc:
        return {"symbol": symbol, "error": f"quote failed: {exc}"}
    if not price > 0:
        return {"symbol": symbol, "error": "no price available"}
    return {"symbol": symbol, "price": round(price, 4)}


def _yfinance_last(symbol: str) -> float:
    import yfinance as yf

    t = yf.Ticker(symbol)
    price = getattr(t.fast_info, "last_price", None)
    if not price:
        hist = t.history(period="1d", interval="1m")
        price = float(hist["Close"].dropna().iloc[-1])
    return float(price)


class _QuoteHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        url = urlparse(self.path)
        if url.path == "/ping":
            body = {"ok": True}
        elif url.path == "/quote":
            symbol = parse_qs(url.query).get("symbol", [""])[0]
            body = quote_payload(symbol)
        else:
            self.send_error(404)
            return
        raw = json.dumps(body).encode()
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        # The lab page opens from file:// (origin "null"); the wildcard is
        # safe because this serves only public quotes, only on loopback.
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def log_message(self, fmt, *fargs):
        pass  # quiet; this runs in a background window


def cmd_serve(args):
    server = ThreadingHTTPServer(("127.0.0.1", args.port), _QuoteHandler)
    print(
        f"Quote helper on http://127.0.0.1:{args.port} — the lab's Refresh button is live."
    )
    print("Ctrl+C stops it.")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass


def main(argv=None):
    ap = argparse.ArgumentParser(
        prog="spicy_lab",
        description="Excel ladder export and quote helper for the spicy lab.",
    )
    sub = ap.add_subparsers(dest="command", required=True)

    p = sub.add_parser("excel", help="write the move ladder workbook for one contract")
    p.add_argument("--symbol", default="?", help="label only")
    p.add_argument("--spot", type=float, required=True)
    p.add_argument("--strike", type=float, required=True)
    p.add_argument(
        "--days", type=float, required=True, help="days to expiry (decimals fine)"
    )
    p.add_argument(
        "--iv", type=float, required=True, help="implied vol in percent, e.g. 16"
    )
    p.add_argument("--kind", default="call", choices=["call", "put"])
    p.add_argument("--premium", type=float, help="actual fill; default = model price")
    p.add_argument("--qty", type=int, default=1)
    p.add_argument(
        "--budget", type=float, default=10, help="decay budget %% for the shot clock"
    )
    p.add_argument("--out", default="spicy_ladder.xlsx")
    p.set_defaults(func=cmd_excel)

    p = sub.add_parser("serve", help="local quote helper for the lab's Refresh button")
    p.add_argument("--port", type=int, default=HELPER_PORT)
    p.set_defaults(func=cmd_serve)

    args = ap.parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
