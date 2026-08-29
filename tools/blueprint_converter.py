#!/usr/bin/env python3
"""
Blueprint Converter – Convert between JSON and Excel formats.

Usage:
    python tools/blueprint_converter.py json-to-xlsx input.json --out output.xlsx
    python tools/blueprint_converter.py xlsx-to-json input.xlsx --out output.json
    python tools/blueprint_converter.py validate input.json
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


STEP_KINDS = ("task", "decision", "delay", "end", "goto")
EXECUTORS = ("person", "automation", "ai")

# A branch reaching further back than this is a long loop-back, and should be a
# go-to step rather than a wire dragged across the map. The rule, and why it
# exists, is in .claude/skills/process-mapping/SKILL.md.
LOOP_BACK_LIMIT = 3

STEP_HEADERS = [
    "Process ID",
    "Number",
    "Title",
    "Kind",
    "Executor",
    "Owner",
    "Duration",
    "Frequency",
    "Tools",
    "Branches",
    "Go To",
    "Notes",
    "Commits",
]


def _as_int(value: Any, fallback: Any = None) -> Any:
    """int(value), or fallback when it is not a whole number."""
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return fallback


def _as_number(value: Any, fallback: Any = None) -> Any:
    """Numeric cell as int where it is whole, else float, else fallback."""
    try:
        n = float(str(value).strip())
    except (TypeError, ValueError):
        return fallback
    return int(n) if n == int(n) else n


def format_branches(branches: Any) -> str:
    """One cell per step: 'Approved > 3; Rejected > end'."""
    parts = []
    for branch in branches or []:
        label = str(branch.get("label", "")).strip()
        parts.append(f"{label} > {branch.get('to', '')}")
    return "; ".join(parts)


def parse_branches(cell: Any) -> list:
    """Read back what format_branches wrote.

    A destination that is neither a step number nor 'end' is kept verbatim so
    validation can name it, rather than being dropped here and going missing.
    """
    out = []
    for chunk in str(cell or "").split(";"):
        chunk = chunk.strip()
        if not chunk or ">" not in chunk:
            continue
        label, _, dest = chunk.rpartition(">")
        label, dest = label.strip(), dest.strip()
        if not label:
            continue
        out.append(
            {"label": label, "to": dest if dest == "end" else _as_int(dest, dest)}
        )
    return out


def step_to_row(process_id: str, step: Dict[str, Any]) -> list:
    """A step as one Steps-sheet row, in STEP_HEADERS order.

    Absent fields stay absent — writing defaults here would invent a `kind` on
    every step of a plain linear process and the round trip would not be one.
    """
    return [
        process_id,
        step.get("number", ""),
        step.get("title", ""),
        step.get("kind", ""),
        step.get("executor", ""),
        step.get("owner", ""),
        step.get("duration", ""),
        step.get("frequency", ""),
        ", ".join(step.get("tools", [])),
        format_branches(step.get("branches")),
        step.get("goto", ""),
        step.get("notes", ""),
        "yes" if step.get("commits") else "",
    ]


def row_to_step(row: Any) -> Dict[str, Any]:
    """One Steps-sheet row back into a step dict."""
    row = tuple(row) + (None,) * (len(STEP_HEADERS) - len(row))
    step: Dict[str, Any] = {
        "number": _as_int(row[1], row[1]),
        "title": row[2] or "",
    }
    for key, cell in (
        ("kind", row[3]),
        ("executor", row[4]),
        ("owner", row[5]),
        ("duration", row[6]),
        ("notes", row[11]),
    ):
        if cell not in (None, ""):
            step[key] = str(cell)
    if row[7] not in (None, ""):
        step["frequency"] = _as_number(row[7], row[7])
    tools = [t.strip() for t in str(row[8] or "").split(",") if t.strip()]
    if tools:
        step["tools"] = tools
    branches = parse_branches(row[9])
    if branches:
        step["branches"] = branches
    if row[10] not in (None, ""):
        step["goto"] = _as_int(row[10], row[10])
    if str(row[12] or "").strip().lower() in ("yes", "true", "1"):
        step["commits"] = True
    return step


def load_json_blueprint(file_path: str) -> Dict[str, Any]:
    """Load a blueprint from JSON file."""
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json_blueprint(blueprint: Dict[str, Any], file_path: str) -> None:
    """Save a blueprint to JSON file."""
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(blueprint, f, indent=2)
    print(f"✅ Saved to {file_path}")


def json_to_xlsx(input_file: str, output_file: str) -> None:
    """Convert JSON blueprint to Excel."""
    if not XLSX_AVAILABLE:
        print("❌ openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    blueprint = load_json_blueprint(input_file)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Metadata sheet
    ws = wb.create_sheet("Metadata")
    ws["A1"] = "Key"
    ws["B1"] = "Value"
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["B1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    ws["B1"].fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )

    meta = blueprint.get("meta", {})
    row = 2
    for key, value in meta.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = str(value)
        row += 1

    # Departments sheet
    ws = wb.create_sheet("Departments")
    headers = ["ID", "Name", "Owner", "Members", "Description", "Processes", "Tools"]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, col).fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )

    for row, dept in enumerate(blueprint.get("departments", []), 2):
        ws.cell(row, 1).value = dept.get("id", "")
        ws.cell(row, 2).value = dept.get("name", "")
        ws.cell(row, 3).value = dept.get("owner", "")
        ws.cell(row, 4).value = dept.get("members", "")
        ws.cell(row, 5).value = dept.get("description", "")
        ws.cell(row, 6).value = ", ".join(dept.get("processes", []))
        ws.cell(row, 7).value = ", ".join(dept.get("tools", []))

    # Processes sheet
    ws = wb.create_sheet("Processes")
    headers = [
        "ID",
        "Name",
        "Category",
        "Owner",
        "Description",
        "Frequency",
        "Metric",
        "Target",
        "Current",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, col).fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )

    for row, proc in enumerate(blueprint.get("processes", []), 2):
        ws.cell(row, 1).value = proc.get("id", "")
        ws.cell(row, 2).value = proc.get("name", "")
        ws.cell(row, 3).value = proc.get("category", "")
        ws.cell(row, 4).value = proc.get("owner", "")
        ws.cell(row, 5).value = proc.get("description", "")
        ws.cell(row, 6).value = proc.get("frequency", "")
        kpi = proc.get("kpi", {})
        ws.cell(row, 7).value = kpi.get("metric", "")
        ws.cell(row, 8).value = kpi.get("target", "")
        ws.cell(row, 9).value = kpi.get("current", "")

    # Steps sheet — the steps of every process, one row each. Without this
    # sheet a round trip through Excel silently deletes every step in the file.
    ws = wb.create_sheet("Steps")
    for col, header in enumerate(STEP_HEADERS, 1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, col).fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )

    row = 2
    for proc in blueprint.get("processes", []):
        for step in proc.get("steps", []):
            for col, value in enumerate(step_to_row(proc.get("id", ""), step), 1):
                ws.cell(row, col).value = value
            row += 1

    # Tools sheet
    ws = wb.create_sheet("Tools")
    headers = [
        "ID",
        "Name",
        "Category",
        "Purpose",
        "Cost",
        "Frequency",
        "Owner",
        "Criticality",
        "Users",
        "Dependencies",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, col).fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )

    for row, tool in enumerate(blueprint.get("tools", []), 2):
        cost = tool.get("cost", {})
        ws.cell(row, 1).value = tool.get("id", "")
        ws.cell(row, 2).value = tool.get("name", "")
        ws.cell(row, 3).value = tool.get("category", "")
        ws.cell(row, 4).value = tool.get("purpose", "")
        ws.cell(row, 5).value = cost.get("amount", "")
        ws.cell(row, 6).value = cost.get("frequency", "")
        ws.cell(row, 7).value = tool.get("owner", "")
        ws.cell(row, 8).value = tool.get("criticality", "")
        ws.cell(row, 9).value = ", ".join(tool.get("users", []))
        ws.cell(row, 10).value = ", ".join(tool.get("dependencies", []))

    # Changes sheet
    ws = wb.create_sheet("Changes")
    headers = ["ID", "Date", "Title", "Category", "Status", "Description", "Author"]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, col).fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )

    for row, change in enumerate(blueprint.get("changes", []), 2):
        ws.cell(row, 1).value = change.get("id", "")
        ws.cell(row, 2).value = change.get("date", "")
        ws.cell(row, 3).value = change.get("title", "")
        ws.cell(row, 4).value = change.get("category", "")
        ws.cell(row, 5).value = change.get("status", "")
        ws.cell(row, 6).value = change.get("description", "")
        ws.cell(row, 7).value = change.get("author", "")

    # Roadmap sheet
    ws = wb.create_sheet("Roadmap")
    headers = [
        "ID",
        "Title",
        "Category",
        "Priority",
        "Target Date",
        "Owner",
        "Status",
        "Description",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, col).fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )

    for row, item in enumerate(blueprint.get("roadmap", []), 2):
        ws.cell(row, 1).value = item.get("id", "")
        ws.cell(row, 2).value = item.get("title", "")
        ws.cell(row, 3).value = item.get("category", "")
        ws.cell(row, 4).value = item.get("priority", "")
        ws.cell(row, 5).value = item.get("targetDate", "")
        ws.cell(row, 6).value = item.get("owner", "")
        ws.cell(row, 7).value = item.get("status", "")
        ws.cell(row, 8).value = item.get("description", "")

    # Auto-adjust column widths
    for ws in wb.sheetnames:
        sheet = wb[ws]
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)
    print(f"✅ Converted to Excel: {output_file}")


def xlsx_to_json(input_file: str, output_file: str) -> None:
    """Convert Excel blueprint to JSON."""
    if not XLSX_AVAILABLE:
        print("❌ openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(input_file)

    blueprint = {
        "meta": {},
        "departments": [],
        "processes": [],
        "tools": [],
        "changes": [],
        "roadmap": [],
    }

    # Read metadata
    if "Metadata" in wb.sheetnames:
        ws = wb["Metadata"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                blueprint["meta"][row[0]] = row[1]

    # Ensure required meta fields
    if "created" not in blueprint["meta"]:
        blueprint["meta"]["created"] = datetime.now().isoformat()
    if "lastModified" not in blueprint["meta"]:
        blueprint["meta"]["lastModified"] = datetime.now().isoformat()

    # Read departments
    if "Departments" in wb.sheetnames:
        ws = wb["Departments"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                blueprint["departments"].append(
                    {
                        "id": row[0] or "",
                        "name": row[1] or "",
                        "owner": row[2] or "",
                        "members": row[3] or 1,
                        "description": row[4] or "",
                        "processes": [
                            p.strip() for p in (row[5] or "").split(",") if p.strip()
                        ],
                        "tools": [
                            t.strip() for t in (row[6] or "").split(",") if t.strip()
                        ],
                    }
                )

    # Read processes
    if "Processes" in wb.sheetnames:
        ws = wb["Processes"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                blueprint["processes"].append(
                    {
                        "id": row[0] or "",
                        "name": row[1] or "",
                        "category": row[2] or "",
                        "owner": row[3] or "",
                        "description": row[4] or "",
                        "frequency": row[5] or "",
                        "steps": [],
                        "kpi": {
                            "metric": row[6] or "",
                            "target": row[7] or "",
                            "current": row[8] or "",
                        },
                    }
                )

    # Read steps back onto their processes
    if "Steps" in wb.sheetnames:
        by_id = {proc["id"]: proc for proc in blueprint["processes"] if proc.get("id")}
        orphans = []
        for row in wb["Steps"].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            proc = by_id.get(row[0])
            if proc is None:
                orphans.append(row[0])
                continue
            proc.setdefault("steps", []).append(row_to_step(row))
        for proc in blueprint["processes"]:
            proc["steps"].sort(key=lambda st: _as_int(st.get("number"), 0))
        if orphans:
            names = ", ".join(sorted(set(orphans)))
            print(f"⚠️  Steps sheet references unknown processes, skipped: {names}")

    # Read tools
    if "Tools" in wb.sheetnames:
        ws = wb["Tools"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                blueprint["tools"].append(
                    {
                        "id": row[0] or "",
                        "name": row[1] or "",
                        "category": row[2] or "",
                        "purpose": row[3] or "",
                        "cost": {
                            "amount": row[4] or 0,
                            "currency": "USD",
                            "frequency": row[5] or "one-time",
                        },
                        "owner": row[6] or "",
                        "criticality": row[7] or "important",
                        "users": [
                            u.strip() for u in (row[8] or "").split(",") if u.strip()
                        ],
                        "dependencies": [
                            d.strip() for d in (row[9] or "").split(",") if d.strip()
                        ],
                    }
                )

    # Read changes
    if "Changes" in wb.sheetnames:
        ws = wb["Changes"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                blueprint["changes"].append(
                    {
                        "id": row[0] or "",
                        "date": str(row[1] or ""),
                        "title": row[2] or "",
                        "category": row[3] or "",
                        "status": row[4] or "completed",
                        "description": row[5] or "",
                        "author": row[6] or "",
                    }
                )

    # Read roadmap
    if "Roadmap" in wb.sheetnames:
        ws = wb["Roadmap"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                blueprint["roadmap"].append(
                    {
                        "id": row[0] or "",
                        "title": row[1] or "",
                        "category": row[2] or "",
                        "priority": row[3] or "medium",
                        "targetDate": str(row[4] or ""),
                        "owner": row[5] or "",
                        "status": row[6] or "backlog",
                        "description": row[7] or "",
                    }
                )

    save_json_blueprint(blueprint, output_file)
    print(f"✅ Converted to JSON: {output_file}")


# A wait, a terminator and a jump are all real parts of a map, and none of them
# is work anybody sits through.
WORK_KINDS = ("task", "decision")


def check_process(proc: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Findings for one process, as {code, severity, step, message}.

    static/process-grammar.js is the same rules for the browser tools, and
    tests/test_process_grammar.py holds the two together by code and step
    number. Adding a rule here means adding it there.
    """
    findings: List[Dict[str, Any]] = []

    def add(code: str, severity: str, step: Any, message: str) -> None:
        findings.append(
            {"code": code, "severity": severity, "step": step, "message": message}
        )

    steps = proc.get("steps", [])
    numbers = [step.get("number") for step in steps]

    seen: Dict[Any, int] = {}
    for number in numbers:
        seen[number] = seen.get(number, 0) + 1
    for number in sorted(n for n, c in seen.items() if c > 1 and n is not None):
        add(
            "duplicate_step_number",
            "error",
            number,
            f"has more than one step numbered {number}",
        )

    known = {n for n in numbers if n is not None}
    unpriced = 0

    for step in steps:
        number = step.get("number")
        kind = step.get("kind", "task")
        executor = step.get("executor", "person")
        branches = step.get("branches") or []

        if kind not in STEP_KINDS:
            add("unknown_kind", "error", number, f"has unknown kind '{kind}'")
        if executor not in EXECUTORS:
            add(
                "unknown_executor",
                "error",
                number,
                f"has unknown executor '{executor}'",
            )

        for branch in branches:
            if not str(branch.get("label", "")).strip():
                add("unlabelled_branch", "error", number, "has a branch with no label")
            dest = branch.get("to")
            if dest == "end":
                continue
            if dest not in known:
                shown = "None" if dest is None else dest
                add(
                    "branch_target_missing",
                    "error",
                    number,
                    f"branches to step {shown}, which does not exist",
                )
            elif (
                isinstance(dest, int)
                and isinstance(number, int)
                and number - dest > LOOP_BACK_LIMIT
            ):
                add(
                    "long_loop_back",
                    "warning",
                    number,
                    f"loops back {number - dest} steps to {dest} — make that a go-to step",
                )

        if kind == "decision" and len(branches) < 2:
            add(
                "thin_fork",
                "warning",
                number,
                "is a decision with fewer than two ways out",
            )
        if kind != "decision" and branches:
            add(
                "branches_on_non_decision",
                "warning",
                number,
                "carries branches but is not a decision",
            )

        if kind == "goto":
            target = step.get("goto")
            if target is None:
                add(
                    "goto_no_destination",
                    "error",
                    number,
                    "is a go-to step with no destination",
                )
            elif target not in known:
                add(
                    "goto_target_missing",
                    "error",
                    number,
                    f"jumps to step {target}, which does not exist",
                )

        # a step with only one of the two numbers cannot be costed, and saying
        # so beats leaving a hole in the total that nothing points at
        if executor == "person" and kind in WORK_KINDS:
            if not step.get("duration") or not step.get("frequency"):
                unpriced += 1

    if unpriced:
        add(
            "unpriced_person_steps",
            "warning",
            None,
            f"{unpriced} person step{'' if unpriced == 1 else 's'} "
            "without both a duration and a monthly frequency",
        )

    return findings


def check_blueprint(blueprint: Dict[str, Any]) -> Tuple[List[str], List[str]]:
    """Return (errors, warnings) for a blueprint. Pure: no printing, no exit."""
    errors: List[str] = []
    warnings: List[str] = []

    # Required fields in meta
    if not blueprint.get("meta", {}).get("name"):
        errors.append("❌ meta.name is required")
    if not blueprint.get("meta", {}).get("owner"):
        errors.append("❌ meta.owner is required")

    # At least one department
    if len(blueprint.get("departments", [])) == 0:
        warnings.append("⚠️  No departments defined")

    # At least one process
    if len(blueprint.get("processes", [])) == 0:
        warnings.append("⚠️  No processes defined")

    # Check all processes have owners
    for proc in blueprint.get("processes", []):
        if not proc.get("owner"):
            errors.append(f"❌ Process '{proc.get('name', '?')}' has no owner")

    # Check all departments have owners
    for dept in blueprint.get("departments", []):
        if not dept.get("owner"):
            errors.append(f"❌ Department '{dept.get('name', '?')}' has no owner")

    # Check all tools have owners
    for tool in blueprint.get("tools", []):
        if not tool.get("owner"):
            errors.append(f"❌ Tool '{tool.get('name', '?')}' has no owner")

    # Steps and branches. A branch that points nowhere is the failure this
    # catches: it reads fine in a list and loses the flow on any map built
    # from it.
    for proc in blueprint.get("processes", []):
        name = proc.get("name") or proc.get("id") or "?"
        for finding in check_process(proc):
            if finding["step"] is None:
                where = f"Process '{name}'"
            else:
                where = f"Process '{name}' step {finding['step']}"
            line = f"{where} {finding['message']}"
            if finding["severity"] == "error":
                errors.append(f"❌ {line}")
            else:
                warnings.append(f"⚠️  {line}")

    return errors, warnings


def validate_blueprint(file_path: str) -> None:
    """Validate a blueprint against the schema."""
    blueprint = load_json_blueprint(file_path)
    errors, warnings = check_blueprint(blueprint)

    # Print results
    if errors:
        print("\n🔴 VALIDATION ERRORS:")
        for error in errors:
            print(error)
        sys.exit(1)

    if warnings:
        print("\n🟡 WARNINGS:")
        for warning in warnings:
            print(warning)

    # Summary
    steps = sum(len(p.get("steps", [])) for p in blueprint.get("processes", []))
    print("\n✅ BLUEPRINT VALID")
    print(f"   Departments: {len(blueprint.get('departments', []))}")
    print(f"   Processes: {len(blueprint.get('processes', []))}")
    print(f"   Process steps: {steps}")
    print(f"   Tools: {len(blueprint.get('tools', []))}")
    print(f"   Changes logged: {len(blueprint.get('changes', []))}")
    print(f"   Roadmap items: {len(blueprint.get('roadmap', []))}")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    command = sys.argv[1]
    input_file = sys.argv[2] if len(sys.argv) > 2 else None
    output_file = None

    if "--out" in sys.argv:
        out_idx = sys.argv.index("--out")
        output_file = sys.argv[out_idx + 1] if out_idx + 1 < len(sys.argv) else None

    if command == "json-to-xlsx":
        if not input_file or not output_file:
            print(
                "Usage: blueprint_converter.py json-to-xlsx input.json --out output.xlsx"
            )
            sys.exit(1)
        json_to_xlsx(input_file, output_file)

    elif command == "xlsx-to-json":
        if not input_file or not output_file:
            print(
                "Usage: blueprint_converter.py xlsx-to-json input.xlsx --out output.json"
            )
            sys.exit(1)
        xlsx_to_json(input_file, output_file)

    elif command == "validate":
        if not input_file:
            print("Usage: blueprint_converter.py validate input.json")
            sys.exit(1)
        validate_blueprint(input_file)

    else:
        print(f"Unknown command: {command}")
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()
