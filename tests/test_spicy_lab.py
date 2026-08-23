"""spicy_lab's ladder math, Excel export, and quote-helper logic — no network."""

import pytest
from openpyxl import load_workbook

from tools.spicy_lab import (
    MINUTES_COLS,
    cmd_excel,
    ladder_rows,
    quote_payload,
    summary_block,
    trading_hours_left,
)

ATM = dict(spot=640.0, strike=640.0, days=1.27, iv=0.16, kind="call")


def test_ladder_up_rungs_pay_and_down_rungs_cost():
    rows = ladder_rows(**ATM)
    by_rung = {r["rung"]: r for r in rows}
    now = lambda r: r["cells"][0]["pnl"]
    assert now(by_rung["+1xEM"]) > 0 > now(by_rung["-1xEM"])
    assert now(by_rung["+2xEM"]) > now(by_rung["+1xEM"])
    # the entry rung at t=0 is exactly flat when premium = model price
    assert now(by_rung["entry"]) == pytest.approx(0.0, abs=0.01)


def test_ladder_time_columns_charge_theta_at_entry_rung():
    rows = ladder_rows(**ATM)
    entry = next(r for r in rows if r["rung"] == "entry")
    pnls = [c["pnl"] for c in entry["cells"]]
    assert pnls == sorted(pnls, reverse=True)  # flat + waiting = strictly worse
    assert pnls[-1] < 0  # two flat hours cost real money


def test_ladder_same_move_pays_less_later():
    rows = ladder_rows(**ATM)
    up = next(r for r in rows if r["rung"] == "+1xEM")
    assert up["cells"][0]["pnl"] > up["cells"][-1]["pnl"]


def test_pct_mode_levels_and_actual_fill_premium():
    rows = ladder_rows(mode="pct", premium=5.0, **ATM)
    one_pct = next(r for r in rows if r["rung"] == "+1%")
    assert one_pct["level"] == pytest.approx(640 * 1.01)
    # entry rung P&L reflects the actual fill, not the model price
    entry = next(r for r in rows if r["rung"] == "entry")
    model_now = entry["cells"][0]["premium"]
    assert entry["cells"][0]["pnl"] == pytest.approx((model_now - 5.0) * 100, abs=0.01)


def test_expired_cells_settle_to_intrinsic():
    rows = ladder_rows(spot=640.0, strike=630.0, days=0.05, iv=0.16, kind="call")
    entry = next(r for r in rows if r["rung"] == "entry")
    # 0.05d = 72 min; the +120m column is past expiry -> intrinsic 10
    assert entry["cells"][-1]["premium"] == pytest.approx(10.0)


def test_summary_block_and_trading_hours():
    s = summary_block(**ATM)
    assert s["delta"] == pytest.approx(0.52, abs=0.05)  # ATM-ish call
    assert s["theta_day"] < 0
    assert s["hourly_hurdle"] > 0
    assert trading_hours_left(0.27) == pytest.approx(6.48, abs=0.01)
    assert trading_hours_left(1.27) == pytest.approx(12.98, abs=0.01)
    assert trading_hours_left(5.0) == pytest.approx(32.5)


def test_excel_export_writes_both_ladders(tmp_path):
    out = tmp_path / "ladder.xlsx"

    class Args:
        symbol = "SPY"
        spot = 640.0
        strike = 640.0
        days = 1.27
        iv = 16.0
        kind = "call"
        premium = None
        qty = 1
        budget = 10.0

    Args.out = str(out)
    cmd_excel(Args)
    ws = load_workbook(out).active
    text = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("Rungs x expected move" in t for t in text)
    assert any("Fixed % rungs" in t for t in text)
    assert any("Shot clock" in t for t in text)
    assert sum(1 for t in text if t == "entry") == 2  # one entry row per ladder


def test_quote_payload_paths():
    assert quote_payload("spy", fetch=lambda s: 641.25) == {
        "symbol": "SPY",
        "price": 641.25,
    }
    assert "error" in quote_payload("", fetch=lambda s: 1.0)
    assert "error" in quote_payload("SPY", fetch=lambda s: 0.0)

    def boom(s):
        raise RuntimeError("feed down")

    out = quote_payload("SPY", fetch=boom)
    assert out["symbol"] == "SPY" and "feed down" in out["error"]
